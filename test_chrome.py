import pychrome
from bs4 import BeautifulSoup
import requests
from awesome import RMPDownload

from openpyxl import load_workbook
from openpyxl import Workbook
import os,sys
import shutil 
import time


FILL_VALUE_IN_FORM = 'document.getElementById("%s").setAttribute("value", "%s");'
CLICK_BY_ID = 'document.getElementById("%s").click()'
#CLICK_ONE_BY_CLASS = 'document.getElementsByClassName("%s").click()'
CLICK_ONE_BY_CLASS = 'document.getElementsByClassName("target")%s.click()'
CLICK_ALL_BY_CLASS = '''
var nodes = document.getElementsByClassName("%s");
for (var i=0;i<nodes.length;i++){
    nodes[i].click();
}
'''
CLICK_ALL_BY_HREF = '''
var nodes = document.querySelectorAll("a[href]");
for (var i=0;i<nodes.length;i++){
    console.log("hello world"); 
    nodes[i].click();
}
'''

CLICK_A_NUM = '''
document.querySelectorAll("a")[%s].click();
'''

class crawler():

    def __init__(self,*args,**kwargs):
        # create a browser instance
        self.browser = pychrome.Browser(url="http://127.0.0.1:9221")
        # create a tab
        self.tab = self.browser.new_tab()
        #self.tab.Network.requestWillBeSent = self.request_will_be_sent
        # start the tab
        self.tab.start()
        # call method
        self.tab.Network.enable()
       
    # register callback if you want
    def request_will_be_sent(**kwargs):
        print("loading: %s" % kwargs.get('request').get('url'))

    def call_web(self,keyword,start,size): 
        list = []
        list2= []
        list3 = []

        # call method with timeout
        self.tab.call_method('Page.navigate',url="https://arxiv.org/search/?query="+keyword+"&searchtype=all&source=header&order=-announced_date_first&size="+size+"&abstracts=show&start="+start, _timeout=20)

        # wait for loading
        self.tab.wait(2)

        html = self.tab.Runtime.evaluate(expression="document.documentElement.outerHTML")

        #print(html)
        soup = BeautifulSoup(((html['result'])['value']),"html.parser")

        for link in soup.find_all("p"):
#            print(link)
#            print(type(str(link)))   
            if str(link).find("title is-5 mathjax") > 0 :
#                print(link) 
                a=str(link).find('''<p class="title is-5 mathjax">''')
                b=str(link).find("</p>")
                title=str(link)[a+46:b-4]
                list.append(title)

#        for i in list:
#            print(i)

        number_pdf=0
        control_1=0

        for link in soup.find_all("a"):
#            print(link)
#           print(type(str(link)))   
            if str(link).find("pdf") > 0 :
                print(link)
                #print(type(str(link)))
                a=str(link).find("href=")
                b=str(link).find("</a>",a)
                addr=str(link)[a+6:b-5]
                a=str(link).find("pdf")
                pdf_name=str(link)[a+4:a+14]
#                print(pdf_name)
                number_pdf=number_pdf+1
                D1={}
                D1["serial_num"]=pdf_name
                D1["addr"]=addr	
 #               print(D1)
                list2.append(D1)

                if(self.if_existence(pdf_name)):
                   control_1=control_1+1
                   list3.append(None)
                   break
                   continue 

                cmd=CLICK_A_NUM % control_1
                print(cmd)
                self.tab.Runtime.evaluate(expression=cmd)

              # wait for loading
                self.tab.wait(20)

                html = self.tab.Runtime.evaluate(expression="document.documentElement.outerHTML")

                #print(html)

                soup = BeautifulSoup(((html['result'])['value']),"html.parser")

                control_2=0
                control_3=0
                for link in soup.find_all("a"):
#                    print(link)
                    if str(link).find("here") > 0 :
                        print(link)
#                        print(type(str(link)))
                        control_3=control_2			                    
                        break
                    control_2=control_2+1

                print("control1 and control_3 ")
                print(control_1)
                print(control_3)
                t=RMPDownload("C:\\Users\\shijonn\\Desktop\\cop\\automation"+"\\")
                t.download_by_quip("https://arxiv.org/search/?query="+keyword+"&searchtype=all&source=header&order=-announced_date_first&size="+size+"&abstracts=show&start="+start,pdf_name+".pdf",control_1,control_3)


                fileList=os.listdir("../")
                for file in fileList:
                    print("file name=",file)
                    if os.path.isdir("../"+file):
                        fileList2=os.listdir("../"+file)
                        if file=="automation":
                            if file2==(pdf_name+".pdf"):
                                list3.append(None)
                            continue    
 
                        for file2 in fileList2:        
                            print("file name=",file2)                                 
                            if file2==(pdf_name+".pdf"):
                                list3.append(file)
                                print(file) 
                                print(type(file)) 
                                print(len(list3))
                                print("there already is the file")
                                os.remove("./"+pdf_name+".pdf")

            control_1=control_1+1

#        if(not((len(list2)==len(list))and(len(list3)==len(list)))):
#            print("something bad happened ")
#            exit()                      

        for i in list2:
            i["title"]=list.pop(0)  
            i["position"]=list3.pop(0)
            print(i['serial_num'],i['addr'],i["title"],i["position"])   
            self.excel("./tamplate/","arxiv.xlsx","./files/","arxiv-total.xlsx",i['serial_num'],i["title"],i['addr'],None,time.asctime( time.localtime(time.time()) ),i["position"])               

    def if_existence(self,file_name): 

        exist=False
        fileList=os.listdir("../")
        for file in fileList:
            if os.path.isdir("../"+file):
                fileList2=os.listdir("../"+file)  
                for file2 in fileList2:        
                    print("file name=",file2)                                 
                    if file2==(file_name+".pdf"):                          
                        exist=True
                        print("there already is the file")
   
        return  exist

    def excel(self,path,name,output,name2,serial_num,titile,addr,existence,date,position):
    
        print(path)
        print(output)

        fileList=os.listdir(output)

        if(len(fileList)==0):
            shutil.copy(path+name,output)   

        fileList=os.listdir(output)

        wkbk = Workbook()

        for file in fileList:
            wb = load_workbook(output+file)
            print(wb.sheetnames[0])

            for sheet in wb:
                print("max row=",sheet.max_row,"max column=",sheet.max_column)
  
                file_num=2
                exist=False
                while(file_num<=sheet.max_row):
                    if(sheet.cell(row=file_num, column=1).value==serial_num):
                        print("is true!")
                        self.record(sheet,file_num,serial_num,titile,addr,existence,date,position)
                        break
                    file_num=file_num+1 

                next_row=sheet.max_row+1

                if(exist==False):
                    self.record(sheet,next_row,serial_num,titile,addr,existence,date,position)
 
                wb.save(output+"tmp.xlsx")
                                
        shutil.copy(output+"tmp.xlsx",output+name2)             
        os.remove(output+"tmp.xlsx")
	
        try:
            os.remove(output+name) 
        except FileNotFoundError: 
            print("ignore it ")


    def record(self,sheet,file_num,serial_num,titile,addr,existence,date,position):

        if(serial_num!=None):
            sheet.cell(row=file_num, column=1,value=serial_num)
        if(titile!=None):
            sheet.cell(row=file_num, column=2,value=titile)
        if(addr!=None):
            sheet.cell(row=file_num, column=3,value=addr)
        if(existence!=None):
            sheet.cell(row=file_num, column=4,value=existence)
        if(date!=None):
            sheet.cell(row=file_num, column=5,value=date)
        if(position!=None):
             sheet.cell(row=file_num, column=6,value=position)



if __name__ == "__main__" :
    c=crawler() 
    c.call_web("Autonomous+Vehicle","0","50")
#    c.call_web("Autonomous+Vehicle","100")
#    c.excel("./files/","./files/","test","test","test","test","test","test","test")












